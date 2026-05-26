#!/usr/bin/env python3
"""
Analyze Excel files from Google Drive backups
Parse structure and identify column mappings
"""

import base64
import io
import sys
import json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

def analyze_excel_file(file_data_base64, filename):
    """Analyze Excel file structure"""
    print(f"\n{'='*60}")
    print(f"Analyzing: {filename}")
    print('='*60)

    try:
        # Decode base64
        file_data = base64.b64decode(file_data_base64)
        file_obj = io.BytesIO(file_data)

        # Load workbook
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active

        print(f"Sheet: {ws.title}")
        print(f"Dimensions: {ws.dimensions}")

        # Get header row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)

        print(f"\nColumns ({len(headers)}):")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")

        # Analyze data rows
        data_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if any(row):  # Skip empty rows
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                data_rows.append(row_dict)

        print(f"\nData Rows: {len(data_rows)}")

        if data_rows:
            print("\nFirst Row Sample:")
            first_row = data_rows[0]
            for key, value in first_row.items():
                print(f"  {key}: {value} ({type(value).__name__})")

        return {
            'filename': filename,
            'sheet': ws.title,
            'columns': headers,
            'row_count': len(data_rows),
            'data_sample': data_rows[0] if data_rows else None,
            'all_rows': data_rows
        }

    except Exception as e:
        print(f"Error analyzing {filename}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    # File data from Google Drive (base64 encoded)
    files = {
        'AdminEmailConfig.xlsx': 'UEsDBBQAAAAAAAAAAACkAYS4tQIAALUCAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHM8P3htbCB2ZXJzaW9uPSIxLjAiIGVuY29kaW5nPSJVVEYtOCIgc3RhbmRhbG9uZT0ieWVzIj8+DQo8UmVsYXRpb25zaGlwcyB4bWxucz0iaHR0cDovL3NjaGVtYXMub3BlbnhtbGZvcm1hdHMub3JnL3BhY2thZ2UvMjAwNi9yZWxhdGlvbnNoaXBzIj48UmVsYXRpb25zaGlwIElkPSJySWQxIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3dvcmtzaGVldCIgVGFyZ2V0PSJ3b3Jrc2hlZXRzL3NoZWV0MS54bWwiLz48UmVsYXRpb25zaGlwIElkPSJySWQyIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3RoZW1lIiBUYXJnZXQ9InRoZW1lL3RoZW1lMS54bWwiLz48UmVsYXRpb25zaGlwIElkPSJySWQzIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3N0eWxlcyIgVGFyZ2V0PSJzdHlsZXMueG1sIi8+PFJlbGF0aW9uc2hpcCBJZD0icklkNCIgVHlwZT0iaHR0cDovL3NjaGVtYXMub3BlbnhtbGZvcm1hdHMub3JnL29mZmljZURvY3VtZW50LzIwMDYvcmVsYXRpb25zaGlwcy9zaGVldE1ldGFkYXRhIiBUYXJnZXQ9Im1ldGFkYXRhLnhtbCIvPjwvUmVsYXRpb25zaGlwcz4',
        'MarketingConfig.xlsx': 'UEsDBBQAAAAAAAAAAACkAYS4tQIAALUCAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHM8P3htbCB2ZXJzaW9uPSIxLjAiIGVuY29kaW5nPSJVVEYtOCIgc3RhbmRhbG9uZT0ieWVzIj8+DQo8UmVsYXRpb25zaGlwcyB4bWxucz0iaHR0cDovL3NjaGVtYXMub3BlbnhtbGZvcm1hdHMub3JnL3BhY2thZ2UvMjAwNi9yZWxhdGlvbnNoaXBzIj48UmVsYXRpb25zaGlwIElkPSJySWQxIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3dvcmtzaGVldCIgVGFyZ2V0PSJ3b3Jrc2hlZXRzL3NoZWV0MS54bWwiLz48UmVsYXRpb25zaGlwIElkPSJySWQyIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3RoZW1lIiBUYXJnZXQ9InRoZW1lL3RoZW1lMS54bWwiLz48UmVsYXRpb25zaGlwIElkPSJySWQzIiBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQvMjAwNi9yZWxhdGlvbnNoaXBzL3N0eWxlcyIgVGFyZ2V0PSJzdHlsZXMueG1sIi8+PFJlbGF0aW9uc2hpcCBJZD0icklkNCIgVHlwZT0iaHR0cDovL3NjaGVtYXMub3BlbnhtbGZvcm1hdHMub3JnL29mZmljZURvY3VtZW50LzIwMDYvcmVsYXRpb25zaGlwcy9zaGVldE1ldGFkYXRhIiBUYXJnZXQ9Im1ldGFkYXRhLnhtbCIvPjwvUmVsYXRpb25zaGlwcz4='
    }

    results = {}
    for filename, base64_data in files.items():
        result = analyze_excel_file(base64_data, filename)
        if result:
            results[filename] = result

    # Print summary report
    print(f"\n\n{'='*60}")
    print("RESTORATION READINESS REPORT")
    print('='*60)

    summary = {
        'timestamp': datetime.now().isoformat(),
        'files_analyzed': len(results),
        'tables_found': []
    }

    for filename, data in results.items():
        if data:
            summary['tables_found'].append({
                'file': filename,
                'sheet': data['sheet'],
                'columns': data['columns'],
                'row_count': data['row_count'],
                'status': 'READY' if data['row_count'] > 0 else 'EMPTY'
            })

    print(json.dumps(summary, indent=2))

    print("\n" + "="*60)
    print("RESTORATION STATUS")
    print("="*60)

    for item in summary['tables_found']:
        status_symbol = '✓' if item['status'] == 'READY' else '✗'
        print(f"{status_symbol} {item['file']:<30} {item['row_count']:>3} rows ready for {item['sheet']}")

if __name__ == '__main__':
    main()
