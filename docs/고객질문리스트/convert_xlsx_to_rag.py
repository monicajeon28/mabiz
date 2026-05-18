#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XLSX → JSON RAG 변환 스크립트
133개의 Q&A를 기존 275개에 추가하여 총 408+개 완성
"""

import json
import pandas as pd
import openpyxl
from datetime import datetime
from typing import List, Dict, Any
import re
import sys
import io

# Windows 터미널 UTF-8 인코딩 설정
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 카테고리 키워드 매핑
CATEGORY_KEYWORDS = {
    "탑승&수속": [
        "탑승", "승선", "수속", "체크인", "여권", "짐", "캐리어", "하선", "수하물",
        "탑승권", "탑승카드", "승선카드", "여권검사", "수속절차", "하선절차",
        "이민", "관세", "짐운반", "포터"
    ],
    "식사&음료": [
        "뷔페", "정찬", "식사", "레스토랑", "조식", "음료", "물", "드링크", "룸서비스",
        "앗치", "카페", "바", "와인", "맥주", "칵테일", "저녁식사", "아침식사",
        "점심", "디너", "브런치", "식사권", "식사포함", "알코올"
    ],
    "객실&카드": [
        "객실", "카드", "신용카드", "냉장고", "미니바", "수건", "비치타올",
        "침대", "욕실", "에어컨", "TV", "핸드폰", "전화", "세이프", "객실문제",
        "객실청소", "룸서비스카드", "부팅카드", "해정키"
    ],
    "선상활동": [
        "공연", "예약", "워터슬라이드", "풀장", "쉽투어", "시설", "극장", "부대시설",
        "스포츠", "피트니스", "스파", "만이", "키즈클럽", "게임", "영화", "쇼",
        "락투어", "스윔", "다이빙", "요가", "댄스", "경기", "콘서트"
    ],
    "기항지&투어": [
        "기항지", "가고시마", "고베", "도쿄", "셔틀버스", "관광", "항구",
        "상해", "홍콩", "싱가포르", "방콕", "푸켓", "제주", "인천", "부산",
        "투어", "관광지", "가이드투어", "선택투어", "버스투어", "육지투어"
    ],
    "정책&수수료": [
        "250달러", "팁", "수수료", "결제", "유료", "무료", "취소", "환불",
        "요금", "가격", "비용", "차이", "왜", "비싼", "비용해설", "팁정책",
        "팁계산", "선상결제", "카드결제", "현금", "달러", "환율"
    ],
    "기술&앱": [
        "와이파이", "어플", "앱", "MSC앱", "콘시어지", "키오스크", "로밍",
        "인터넷", "와이파이비용", "핫스팟", "휴대폰", "신호", "통신", "전자기기",
        "카메라", "배터리", "충전기", "어댑터", "110V", "220V"
    ]
}


def get_category(question: str, answer: str = "") -> str:
    """질문과 답변으로부터 카테고리 결정"""
    text = (question + " " + answer).lower()

    # 카테고리별 키워드 점수 계산
    scores = {}
    for category, keywords in CATEGORY_KEYWORDS.items():
        score = sum(1 for keyword in keywords if keyword in text)
        if score > 0:
            scores[category] = score

    # 최고 점수의 카테고리 반환
    if scores:
        return max(scores, key=scores.get)
    return "기타"


def generate_keywords(text: str, category: str) -> List[str]:
    """텍스트로부터 키워드 추출"""
    words = text.split()
    # 카테고리 키워드도 포함
    base_keywords = CATEGORY_KEYWORDS.get(category, [])

    # 길이가 2 이상인 단어들 추출
    extracted = [w for w in words if len(w) >= 2][:5]

    # 기본 키워드와 추출 키워드 합치기
    keywords = list(set(base_keywords[:3] + extracted))
    return keywords[:5]


def process_sheet1(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트1: 안내문제 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[1]:  # col[1]이 비어있으면 스킵
            continue

        try:
            question = f"[실전문제] {row[1]}"
            # 답변은 col[2]에서 가져오거나, 없으면 자동 생성
            answer = row[2] if len(row) > 2 and row[2] else f"이것은 MSC 크루즈 탑승 관련 문제입니다."

            item = {
                "id": f"msc2605_1_{seq:04d}",
                "key": f"msc2605_1_{seq:04d}",
                "question": str(question),
                "answer": str(answer),
                "category": get_category(question, answer),
                "source": "MSC 2026-05 교육자료",
                "type": "qa",
                "salesTone": "friendly",
                "keywords": generate_keywords(question, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트1 행{row_idx}: {e}")
            continue

    return items


def process_sheet2(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트2: Q&A 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[1]:
            continue

        try:
            question = str(row[1])
            answer = str(row[2]) if len(row) > 2 and row[2] else ""

            if not answer:
                continue

            item = {
                "id": f"msc2605_2_{seq:04d}",
                "key": f"msc2605_2_{seq:04d}",
                "question": question,
                "answer": answer,
                "category": get_category(question, answer),
                "source": "MSC 2026-05 Q&A",
                "type": "qa",
                "salesTone": "friendly",
                "keywords": generate_keywords(question, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트2 행{row_idx}: {e}")
            continue

    return items


def process_sheet3(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트3: 꿀팁 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[1]:
            continue

        try:
            tip_text = str(row[1])
            # 첫 50자로 질문 생성
            question = f"[꿀팁] {tip_text[:50]}?" if len(tip_text) > 50 else f"[꿀팁] {tip_text}?"
            answer = tip_text

            item = {
                "id": f"msc2605_3_{seq:04d}",
                "key": f"msc2605_3_{seq:04d}",
                "question": question,
                "answer": answer,
                "category": get_category(question, answer),
                "source": "MSC 크루즈 꿀팁",
                "type": "tip",
                "salesTone": "friendly",
                "keywords": generate_keywords(tip_text, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트3 행{row_idx}: {e}")
            continue

    return items


def process_sheet4(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트4: 건의 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[1]:
            continue

        try:
            suggestion = str(row[1])
            author = str(row[2]) if len(row) > 2 and row[2] else "익명"

            question = f"[건의] {suggestion[:50]}?"
            answer = f"{suggestion} (작성자: {author})"

            item = {
                "id": f"msc2605_4_{seq:04d}",
                "key": f"msc2605_4_{seq:04d}",
                "question": question,
                "answer": answer,
                "category": get_category(question, answer),
                "source": "고객 건의사항",
                "type": "feedback",
                "salesTone": "professional",
                "keywords": generate_keywords(suggestion, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트4 행{row_idx}: {e}")
            continue

    return items


def process_sheet5(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트5: 공지사항 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[0]:
            continue

        try:
            question = str(row[0])
            answer = str(row[1]) if len(row) > 1 and row[1] else ""

            if not answer:
                continue

            item = {
                "id": f"msc2605_5_{seq:04d}",
                "key": f"msc2605_5_{seq:04d}",
                "question": question,
                "answer": answer,
                "category": get_category(question, answer),
                "source": "MSC 공지사항",
                "type": "notice",
                "salesTone": "professional",
                "keywords": generate_keywords(question, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트5 행{row_idx}: {e}")
            continue

    return items


def process_sheet6(ws, start_row: int = 2) -> List[Dict[str, Any]]:
    """시트6: 스탭 준비 처리"""
    items = []
    seq = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row or not row[1]:
            continue

        try:
            prep_text = str(row[1])
            detail = str(row[2]) if len(row) > 2 and row[2] else ""

            question = f"[스탭준비] {prep_text}"
            answer = detail if detail else prep_text

            item = {
                "id": f"msc2605_6_{seq:04d}",
                "key": f"msc2605_6_{seq:04d}",
                "question": question,
                "answer": answer,
                "category": get_category(question, answer),
                "source": "스탭 준비 가이드",
                "type": "guide",
                "salesTone": "professional",
                "keywords": generate_keywords(prep_text, get_category(question, answer)),
                "isActive": True,
                "createdAt": datetime.now().isoformat()
            }
            items.append(item)
            seq += 1
        except Exception as e:
            print(f"  [경고] 시트6 행{row_idx}: {e}")
            continue

    return items


def main():
    input_file = "docs/고객질문리스트/2026년5월 MSC 크루즈 여행 정보, 질문 기록.xlsx"
    output_file = "docs/고객질문리스트/msc_2026_05_qa.json"

    print("=" * 60)
    print("XLSX -> JSON RAG Conversion Start")
    print("=" * 60)
    print("Input: {}".format(input_file))
    print("Output: {}".format(output_file))
    print()

    try:
        # 엑셀 파일 로드
        wb = openpyxl.load_workbook(input_file)
        sheet_names = wb.sheetnames

        print("Total {} sheets found:".format(len(sheet_names)))
        for i, name in enumerate(sheet_names, 1):
            ws = wb[name]
            print("  {}. {} ({} rows)".format(i, name, ws.max_row))
        print()

        all_items = []

        # 각 시트별 처리
        for i, sheet_name in enumerate(sheet_names, 1):
            ws = wb[sheet_name]
            print("[Sheet{}] {} processing...".format(i, sheet_name), end=" ")

            try:
                if i == 1:
                    items = process_sheet1(ws)
                elif i == 2:
                    items = process_sheet2(ws)
                elif i == 3:
                    items = process_sheet3(ws)
                elif i == 4:
                    items = process_sheet4(ws)
                elif i == 5:
                    items = process_sheet5(ws)
                elif i == 6:
                    items = process_sheet6(ws)
                else:
                    items = []

                print("[OK] {} items created".format(len(items)))
                all_items.extend(items)

            except Exception as e:
                print("[ERROR] {}".format(e))
                continue

        print()
        print("=" * 60)
        print("Total {} items created".format(len(all_items)))
        print("=" * 60)
        print()

        # 중복 체크 (key 기준)
        keys = [item['key'] for item in all_items]
        duplicates = len(keys) - len(set(keys))
        if duplicates > 0:
            print("[WARNING] {} duplicate keys found!".format(duplicates))
        else:
            print("[OK] No duplicate keys")

        # JSON 저장
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump({
                "total": len(all_items),
                "items": all_items,
                "metadata": {
                    "source": "MSC 2026-05 Travel Information",
                    "convertedAt": datetime.now().isoformat(),
                    "sheetCount": len(sheet_names)
                }
            }, f, ensure_ascii=False, indent=2)

        print("[OK] JSON file saved: {}".format(output_file))
        print()

        # 파일 유효성 검사
        with open(output_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        print("=" * 60)
        print("File Validation")
        print("=" * 60)
        print("[OK] JSON validation passed")
        print("[OK] Total items: {}".format(data['total']))
        print("[OK] Category distribution:")

        categories = {}
        for item in data['items']:
            cat = item['category']
            categories[cat] = categories.get(cat, 0) + 1

        for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            print("    - {}: {} items".format(cat, count))

        print()
        print("[OK] Type distribution:")
        types = {}
        for item in data['items']:
            t = item.get('type', 'unknown')
            types[t] = types.get(t, 0) + 1

        for t, count in sorted(types.items()):
            print("    - {}: {} items".format(t, count))

        print()
        print("=" * 60)
        print("[DONE] Conversion completed!")
        print("=" * 60)
        print()
        print("Next step: Run upload_msc_qa.py")

    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
